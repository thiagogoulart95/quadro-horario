import openpyxl
from tabulate import tabulate

# Mapeamento de números para os dias da semana
DAY_MAP = {
    2: "SEGUNDA-FEIRA",
    3: "TERÇA-FEIRA",
    4: "QUARTA-FEIRA",
    5: "QUINTA-FEIRA",
    6: "SEXTA-FEIRA",
}

def ler_planilha(nome_arquivo):
    """
    Lê o arquivo XLSX e retorna:
      - Nome do professor (presumindo que seja o mesmo em todos os registros)
      - Um dicionário de horários agrupado por faixa (hora inicial, hora final)
    """
    quadro = {}
    nome_professor = None
    dias = ["SEGUNDA-FEIRA", "TERÇA-FEIRA", "QUARTA-FEIRA", "QUINTA-FEIRA", "SEXTA-FEIRA"]

    # Abre o arquivo Excel
    wb = openpyxl.load_workbook(nome_arquivo, data_only=True)
    sheet = wb.active  # Usa a primeira aba

    # Lê as colunas (assumindo que a primeira linha tem os cabeçalhos)
    colunas = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}

    # Verifica se todas as colunas necessárias existem
    colunas_necessarias = ["PROFESSOR", "DIASEMANA", "HORAINICIAL", "HORAFINAL", "DISCIPLINA", "CODTURMA"]
    for coluna in colunas_necessarias:
        if coluna not in colunas:
            raise ValueError(f"Erro: A coluna '{coluna}' não foi encontrada na planilha.")

    # Percorre as linhas a partir da segunda (os dados começam aqui)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        linha = {chave: row[colunas[chave] - 1] for chave in colunas_necessarias}

        if not nome_professor:
            nome_professor = linha["PROFESSOR"]

        # Define a faixa horária
        faixa = (linha["HORAINICIAL"], linha["HORAFINAL"])
        dia_nome = DAY_MAP.get(linha["DIASEMANA"], "DIA_INVÁLIDO")
        info = (linha["DISCIPLINA"], linha["CODTURMA"])

        if faixa not in quadro:
            quadro[faixa] = {dia: "" for dia in dias}

        quadro[faixa][dia_nome] = info

    return nome_professor, quadro

def montar_tabela_tabulate(nome_professor, quadro):
    """
    Monta e exibe o quadro de horário formatado.
    """
    dias = ["SEGUNDA-FEIRA", "TERÇA-FEIRA", "QUARTA-FEIRA", "QUINTA-FEIRA", "SEXTA-FEIRA"]
    cabecalho = ["Horário"] + dias
    table_rows = []

    for faixa in sorted(quadro.keys(), key=lambda x: x[0]):
        hora_ini, hora_fin = faixa
        row_disciplina = [hora_ini]
        row_turma = [hora_fin]

        for dia in dias:
            conteudo = quadro[faixa].get(dia, "")
            if conteudo:
                disciplina, turma = conteudo
                row_disciplina.append(disciplina)
                row_turma.append(turma)
            else:
                row_disciplina.append("")
                row_turma.append("")

        table_rows.append(row_disciplina)
        table_rows.append(row_turma)

    print(nome_professor.upper())
    print(tabulate(table_rows, headers=cabecalho, tablefmt="grid"))

    return cabecalho, table_rows

def salvar_em_xlsx(nome_professor, cabecalho, table_rows, nome_arquivo="resultado.xlsx"):
    """
    Salva a tabela formatada em um novo arquivo Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horário"

    # Adiciona o nome do professor na célula A1
    ws.append([nome_professor.upper()])
    ws.append([])  # Linha vazia para separação

    # Adiciona o cabeçalho
    ws.append(cabecalho)

    # Adiciona as linhas de dados
    for row in table_rows:
        ws.append(row)

    # Ajusta a largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Margem extra

    # Salva o arquivo
    wb.save(nome_arquivo)
    print(f"\n📂 Planilha salva como: {nome_arquivo}")

if __name__ == "__main__":
    nome_arquivo = r"C:\Users\thiago.goulart\Desktop\Horário\horario.xlsx"
    professor, quadro_horario = ler_planilha(nome_arquivo)
    cabecalho, table_rows = montar_tabela_tabulate(professor, quadro_horario)
    salvar_em_xlsx(professor, cabecalho, table_rows)
